#!/usr/bin/env python3
"""
Kubernetes Service Health Monitoring & Alert System
Checks readiness and liveness probes for services across multiple clusters
Dynamically fetches namespaces from each cluster
"""

import os
import sys
import yaml
import json
import requests
import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, asdict
from kubernetes import client, config
from kubernetes.client.rest import ApiException
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import logging
from logging.handlers import RotatingFileHandler
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        RotatingFileHandler('service_monitor.log', maxBytes=10485760, backupCount=5),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class ServiceProbe:
    """Data class for service probe configuration"""
    service_name: str
    namespace: str
    cluster: str
    readiness_path: str
    readiness_port: int
    liveness_path: str
    liveness_port: int
    pod_ip: str = None
    node_name: str = None
    status: str = "Unknown"
    last_check: datetime = None
    error_message: str = None
    service_type: str = "Deployment"  # Deployment, StatefulSet, DaemonSet

@dataclass
class ClusterConfig:
    """Data class for cluster configuration"""
    name: str
    context: str
    api_endpoint: str = None
    token: str = None
    ca_cert: str = None
    exclude_namespaces: List[str] = None
    include_only_namespaces: List[str] = None

class KubernetesServiceMonitor:
    """Main class for monitoring Kubernetes services with dynamic namespace discovery"""
    
    def __init__(self, config_file: str = "cluster_config.yaml"):
        """
        Initialize the monitor with cluster configurations
        
        Args:
            config_file: Path to cluster configuration file
        """
        self.clusters: List[ClusterConfig] = []
        self.service_probes: List[ServiceProbe] = []
        self.results: List[Dict] = []
        self.failed_services: List[Dict] = []
        
        self.load_cluster_config(config_file)
        self.setup_email_config()
        
        # Default namespaces to exclude (system namespaces)
        self.default_exclude_namespaces = [
            'kube-system',
            'kube-public',
            'kube-node-lease',
            'default',
            'ingress-nginx',
            'cert-manager',
            'monitoring',
            'logging',
            'tracing',
            'istio-system',
            'kiali-operator',
            'jaeger',
            'grafana',
            'prometheus'
        ]
        
        logger.info(f"Initialized monitor with {len(self.clusters)} clusters")
    
    def load_cluster_config(self, config_file: str):
        """Load cluster configurations from YAML file"""
        try:
            with open(config_file, 'r') as f:
                config_data = yaml.safe_load(f)
                
            for cluster_data in config_data.get('clusters', []):
                cluster = ClusterConfig(
                    name=cluster_data['name'],
                    context=cluster_data.get('context', cluster_data['name'].lower().replace('-', '_')),
                    api_endpoint=cluster_data.get('api_endpoint'),
                    token=cluster_data.get('token'),
                    ca_cert=cluster_data.get('ca_cert'),
                    exclude_namespaces=cluster_data.get('exclude_namespaces', []),
                    include_only_namespaces=cluster_data.get('include_only_namespaces')
                )
                self.clusters.append(cluster)
                
            logger.info(f"Loaded configurations for {len(self.clusters)} clusters")
            
        except FileNotFoundError:
            logger.error(f"Configuration file {config_file} not found")
            # Create default config
            self.create_default_config()
            logger.info("Created default configuration file. Please update it with your cluster contexts.")
            sys.exit(1)
        except Exception as e:
            logger.error(f"Error loading configuration: {str(e)}")
            sys.exit(1)
    
    def create_default_config(self):
        """Create default configuration file"""
        default_config = {
            'clusters': [
                {
                    'name': 'UAT-General',
                    'context': 'uat-general',
                    'exclude_namespaces': ['kube-system', 'kube-public', 'default']
                },
                {
                    'name': 'UAT-BZ-General',
                    'context': 'uatbz-general',
                    'exclude_namespaces': ['kube-system', 'kube-public', 'default']
                },
                {
                    'name': 'UAT-BZ-AI',
                    'context': 'uatbz-ai',
                    'exclude_namespaces': ['kube-system', 'kube-public', 'default']
                },
                {
                    'name': 'UAT-BZ-Bot',
                    'context': 'uatbz-bot',
                    'exclude_namespaces': ['kube-system', 'kube-public', 'default']
                },
                {
                    'name': 'Preprod-General',
                    'context': 'preprod-general',
                    'exclude_namespaces': ['kube-system', 'kube-public', 'default']
                },
                {
                    'name': 'Preprod-AI',
                    'context': 'preprod-ai',
                    'exclude_namespaces': ['kube-system', 'kube-public', 'default']
                },
                {
                    'name': 'Preprod-Bot',
                    'context': 'preprod-bot',
                    'exclude_namespaces': ['kube-system', 'kube-public', 'default']
                }
            ],
            'settings': {
                'check_interval': 300,  # 5 minutes
                'timeout': 10,
                'retry_count': 3
            }
        }
        
        with open('cluster_config.yaml', 'w') as f:
            yaml.dump(default_config, f, default_flow_style=False)
    
    def setup_email_config(self):
        """Setup email configuration"""
        self.email_config = {
            'smtp_server': os.getenv('SMTP_SERVER', 'smtp.gmail.com'),
            'smtp_port': int(os.getenv('SMTP_PORT', 587)),
            'sender_email': os.getenv('SENDER_EMAIL', 'alerts@yourcompany.com'),
            'sender_password': os.getenv('SENDER_PASSWORD', ''),
            'recipients': os.getenv('RECIPIENTS', 'devops-team@yourcompany.com').split(','),
            'cc_recipients': os.getenv('CC_RECIPIENTS', 'managers@yourcompany.com').split(',')
        }
    
    def get_kube_client(self, cluster: ClusterConfig):
        """
        Get Kubernetes client for specific cluster context
        
        Args:
            cluster: Cluster configuration object
        
        Returns:
            Tuple of (v1_client, apps_v1_client)
        """
        try:
            # Try to load from kubeconfig first
            try:
                config.load_kube_config(context=cluster.context)
                logger.debug(f"Loaded kubeconfig for cluster: {cluster.name}")
            except Exception:
                # Fallback to in-cluster config or direct API
                if cluster.api_endpoint and cluster.token:
                    configuration = client.Configuration()
                    configuration.host = cluster.api_endpoint
                    configuration.api_key = {"authorization": f"Bearer {cluster.token}"}
                    if cluster.ca_cert:
                        configuration.ssl_ca_cert = cluster.ca_cert
                    client.Configuration.set_default(configuration)
                    logger.debug(f"Using direct API for cluster: {cluster.name}")
                else:
                    # Try in-cluster config
                    config.load_incluster_config()
                    logger.debug(f"Using in-cluster config for: {cluster.name}")
            
            # Create API clients
            v1 = client.CoreV1Api()
            apps_v1 = client.AppsV1Api()
            
            return v1, apps_v1
            
        except Exception as e:
            logger.error(f"Error creating kube client for {cluster.name}: {str(e)}")
            return None, None
    
    def fetch_namespaces_for_cluster(self, cluster: ClusterConfig) -> List[str]:
        """
        Dynamically fetch namespaces from cluster
        
        Args:
            cluster: Cluster configuration object
        
        Returns:
            List of namespace names
        """
        v1, _ = self.get_kube_client(cluster)
        if not v1:
            logger.error(f"Cannot connect to cluster {cluster.name}")
            return []
        
        try:
            # Get all namespaces
            namespaces = v1.list_namespace()
            namespace_names = [ns.metadata.name for ns in namespaces.items]
            
            logger.info(f"Cluster {cluster.name} has {len(namespace_names)} total namespaces")
            
            # Apply filtering
            filtered_namespaces = []
            
            for ns in namespace_names:
                # Check if namespace should be included
                if cluster.include_only_namespaces:
                    if ns in cluster.include_only_namespaces:
                        filtered_namespaces.append(ns)
                else:
                    # Check if namespace should be excluded
                    exclude_list = self.default_exclude_namespaces + (cluster.exclude_namespaces or [])
                    if ns not in exclude_list:
                        filtered_namespaces.append(ns)
            
            logger.info(f"Cluster {cluster.name}: {len(filtered_namespaces)} namespaces after filtering")
            
            if filtered_namespaces:
                logger.debug(f"Namespaces for {cluster.name}: {', '.join(filtered_namespaces[:10])}" + 
                           ("..." if len(filtered_namespaces) > 10 else ""))
            
            return filtered_namespaces
            
        except ApiException as e:
            logger.error(f"API error fetching namespaces for {cluster.name}: {str(e)}")
            return []
        except Exception as e:
            logger.error(f"Error fetching namespaces for {cluster.name}: {str(e)}")
            return []
    
    def discover_service_probes(self):
        """Discover all deployments and their probes across all clusters"""
        logger.info("Starting service probe discovery...")
        
        total_services = 0
        
        for cluster in self.clusters:
            logger.info(f"Discovering services in cluster: {cluster.name}")
            
            # Dynamically fetch namespaces for this cluster
            namespaces = self.fetch_namespaces_for_cluster(cluster)
            
            if not namespaces:
                logger.warning(f"No namespaces found for cluster {cluster.name}")
                continue
            
            v1, apps_v1 = self.get_kube_client(cluster)
            if not v1 or not apps_v1:
                continue
            
            cluster_services = 0
            
            for namespace in namespaces:
                logger.info(f"  Checking namespace: {namespace}")
                
                try:
                    # Get workload resources
                    deployments = apps_v1.list_namespaced_deployment(namespace=namespace)
                    statefulsets = apps_v1.list_namespaced_stateful_set(namespace=namespace)
                    daemonsets = apps_v1.list_namespaced_daemon_set(namespace=namespace)
                    
                    # Process deployments
                    for deployment in deployments.items:
                        self.extract_probe_from_workload(
                            deployment, 
                            namespace, 
                            cluster, 
                            "Deployment",
                            deployment.spec.template if deployment.spec else None
                        )
                        cluster_services += 1
                    
                    # Process statefulsets
                    for statefulset in statefulsets.items:
                        self.extract_probe_from_workload(
                            statefulset,
                            namespace,
                            cluster,
                            "StatefulSet",
                            statefulset.spec.template if statefulset.spec else None
                        )
                        cluster_services += 1
                    
                    # Process daemonsets
                    for daemonset in daemonsets.items:
                        self.extract_probe_from_workload(
                            daemonset,
                            namespace,
                            cluster,
                            "DaemonSet",
                            daemonset.spec.template if daemonset.spec else None
                        )
                        cluster_services += 1
                    
                    logger.debug(f"    Found {len(deployments.items)} deployments, "
                               f"{len(statefulsets.items)} statefulsets, "
                               f"{len(daemonsets.items)} daemonsets in {namespace}")
                    
                except ApiException as e:
                    logger.error(f"API error in namespace {namespace}: {str(e)}")
                    continue
                except Exception as e:
                    logger.error(f"Error processing namespace {namespace}: {str(e)}")
                    continue
            
            total_services += cluster_services
            logger.info(f"  Cluster {cluster.name}: Found {cluster_services} services")
        
        logger.info(f"Total discovered services: {total_services}")
        logger.info(f"Services with probes: {len(self.service_probes)}")
    
    def extract_probe_from_workload(self, workload, namespace: str, cluster: ClusterConfig, 
                                  workload_type: str, template):
        """Extract probe configuration from workload spec"""
        try:
            service_name = workload.metadata.name
            
            if not template or not template.spec or not template.spec.containers:
                logger.debug(f"    Skipping {workload_type}/{service_name}: No containers found")
                return
            
            # We'll check all containers in the pod
            for container in template.spec.containers:
                # Extract readiness probe
                readiness_probe = container.readiness_probe
                readiness_path = None
                readiness_port = None
                
                if readiness_probe and hasattr(readiness_probe, 'http_get') and readiness_probe.http_get:
                    readiness_path = readiness_probe.http_get.path
                    readiness_port = readiness_probe.http_get.port
                    
                    # Handle named ports
                    if isinstance(readiness_port, str):
                        readiness_port = self.resolve_named_port(container.ports, readiness_port)
                
                # Extract liveness probe
                liveness_probe = container.liveness_probe
                liveness_path = None
                liveness_port = None
                
                if liveness_probe and hasattr(liveness_probe, 'http_get') and liveness_probe.http_get:
                    liveness_path = liveness_probe.http_get.path
                    liveness_port = liveness_probe.http_get.port
                    
                    if isinstance(liveness_port, str):
                        liveness_port = self.resolve_named_port(container.ports, liveness_port)
                
                # If either probe exists, create a service probe entry
                if readiness_path or liveness_path:
                    # Get pod IP for the workload
                    pod_ip, node_name = self.get_pod_info_for_workload(
                        namespace, service_name, cluster, workload_type
                    )
                    
                    probe = ServiceProbe(
                        service_name=service_name,
                        namespace=namespace,
                        cluster=cluster.name,
                        readiness_path=readiness_path,
                        readiness_port=readiness_port,
                        liveness_path=liveness_path,
                        liveness_port=liveness_port,
                        pod_ip=pod_ip,
                        node_name=node_name,
                        service_type=workload_type
                    )
                    
                    self.service_probes.append(probe)
                    logger.debug(f"    Found probes for {workload_type}/{service_name} in container {container.name}")
                    break  # Only need to check first container with probes
        
        except Exception as e:
            logger.warning(f"Error extracting probes for {workload_type}/{service_name}: {str(e)}")
    
    def resolve_named_port(self, container_ports, port_name: str) -> Optional[int]:
        """Resolve named port to container port number"""
        if not container_ports:
            return None
        
        for port in container_ports:
            if port.name == port_name:
                return port.container_port
        
        return None
    
    def get_pod_info_for_workload(self, namespace: str, workload_name: str, 
                                cluster: ClusterConfig, workload_type: str) -> Tuple[Optional[str], Optional[str]]:
        """Get pod IP and node name for a workload"""
        try:
            v1, _ = self.get_kube_client(cluster)
            if not v1:
                return None, None
            
            # Different label selectors for different workload types
            if workload_type == "Deployment":
                label_selector = f"app={workload_name}"
            elif workload_type == "StatefulSet":
                label_selector = f"app={workload_name}"
            elif workload_type == "DaemonSet":
                label_selector = f"app={workload_name}"
            else:
                label_selector = f"app={workload_name}"
            
            pods = v1.list_namespaced_pod(namespace=namespace, label_selector=label_selector)
            
            if pods.items:
                pod = pods.items[0]  # Get first pod
                return pod.status.pod_ip, pod.spec.node_name
            
            # Try alternative label selectors
            alternative_selectors = [
                f"k8s-app={workload_name}",
                f"name={workload_name}",
                f"component={workload_name}",
                workload_name  # Some pods have the same name as the workload
            ]
            
            for selector in alternative_selectors:
                pods = v1.list_namespaced_pod(namespace=namespace, label_selector=selector)
                if pods.items:
                    pod = pods.items[0]
                    return pod.status.pod_ip, pod.spec.node_name
                
        except Exception as e:
            logger.debug(f"Error getting pod info for {workload_name}: {str(e)}")
        
        return None, None
    
    def check_service_probes(self):
        """Check all discovered service probes"""
        logger.info("Starting probe checks...")
        
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import concurrent.futures
        
        # Use thread pool for parallel checking
        max_workers = min(50, len(self.service_probes))
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all probe checks
            future_to_probe = {
                executor.submit(self.check_single_probe, probe): probe 
                for probe in self.service_probes
            }
            
            # Collect results as they complete
            completed = 0
            for future in as_completed(future_to_probe):
                probe = future_to_probe[future]
                try:
                    result = future.result(timeout=30)
                    self.results.append(result)
                    
                    if result['overall_status'] in ['Unhealthy', 'Error', 'Timeout', 'Connection Error']:
                        self.failed_services.append(result)
                    
                    completed += 1
                    if completed % 10 == 0:
                        logger.info(f"Completed {completed}/{len(self.service_probes)} checks")
                        
                except concurrent.futures.TimeoutError:
                    logger.error(f"Timeout checking {probe.service_name}")
                    result = self.create_error_result(probe, "Check timeout")
                    self.results.append(result)
                    self.failed_services.append(result)
                except Exception as e:
                    logger.error(f"Error checking {probe.service_name}: {str(e)}")
                    result = self.create_error_result(probe, str(e))
                    self.results.append(result)
                    self.failed_services.append(result)
        
        logger.info(f"Completed all probe checks. Failed services: {len(self.failed_services)}")
    
    def check_single_probe(self, probe: ServiceProbe) -> Dict:
        """Check a single service probe"""
        result = {
            'timestamp': datetime.now().isoformat(),
            'cluster': probe.cluster,
            'namespace': probe.namespace,
            'service': probe.service_name,
            'service_type': probe.service_type,
            'pod_ip': probe.pod_ip,
            'node_name': probe.node_name,
            'readiness_status': 'Not Configured',
            'liveness_status': 'Not Configured',
            'readiness_response_time': None,
            'liveness_response_time': None,
            'readiness_path': probe.readiness_path,
            'liveness_path': probe.liveness_path,
            'readiness_port': probe.readiness_port,
            'liveness_port': probe.liveness_port,
            'overall_status': 'Unknown',
            'error': None
        }
        
        try:
            # Check readiness probe
            if probe.readiness_path and probe.readiness_port and probe.pod_ip:
                readiness_status, readiness_time = self.check_probe_endpoint(
                    probe.pod_ip, probe.readiness_port, probe.readiness_path
                )
                result['readiness_status'] = readiness_status
                result['readiness_response_time'] = readiness_time
            
            # Check liveness probe
            if probe.liveness_path and probe.liveness_port and probe.pod_ip:
                liveness_status, liveness_time = self.check_probe_endpoint(
                    probe.pod_ip, probe.liveness_port, probe.liveness_path
                )
                result['liveness_status'] = liveness_status
                result['liveness_response_time'] = liveness_time
            
            # Determine overall status
            result['overall_status'] = self.determine_overall_status(result)
            
        except Exception as e:
            result['error'] = str(e)
            result['overall_status'] = 'Error'
        
        # Update probe status
        probe.status = result['overall_status']
        probe.last_check = datetime.now()
        probe.error_message = result.get('error')
        
        return result
    
    def create_error_result(self, probe: ServiceProbe, error_message: str) -> Dict:
        """Create error result for a failed check"""
        return {
            'timestamp': datetime.now().isoformat(),
            'cluster': probe.cluster,
            'namespace': probe.namespace,
            'service': probe.service_name,
            'service_type': probe.service_type,
            'pod_ip': probe.pod_ip,
            'node_name': probe.node_name,
            'readiness_status': 'Error',
            'liveness_status': 'Error',
            'readiness_response_time': None,
            'liveness_response_time': None,
            'overall_status': 'Error',
            'error': error_message
        }
    
    def check_probe_endpoint(self, pod_ip: str, port: int, path: str) -> Tuple[str, float]:
        """
        Check a single probe endpoint
        
        Args:
            pod_ip: Pod IP address
            port: Port number
            path: HTTP path
            
        Returns:
            Tuple of (status, response_time)
        """
        try:
            url = f"http://{pod_ip}:{port}{path}"
            
            start_time = datetime.now()
            response = requests.get(
                url,
                timeout=10,
                headers={
                    'User-Agent': 'K8s-Monitor/1.0',
                    'Accept': 'application/json',
                    'Connection': 'close'
                },
                verify=False
            )
            response_time = (datetime.now() - start_time).total_seconds() * 1000  # Convert to ms
            
            if 200 <= response.status_code < 300:
                return 'Healthy', round(response_time, 2)
            else:
                return f'HTTP {response.status_code}', round(response_time, 2)
                
        except requests.exceptions.Timeout:
            return 'Timeout', 10000
        except requests.exceptions.ConnectionError:
            return 'Connection Error', None
        except requests.exceptions.SSLError:
            # Try HTTPS if HTTP failed
            try:
                url = f"https://{pod_ip}:{port}{path}"
                start_time = datetime.now()
                response = requests.get(url, timeout=5, verify=False)
                response_time = (datetime.now() - start_time).total_seconds() * 1000
                if 200 <= response.status_code < 300:
                    return 'Healthy (HTTPS)', round(response_time, 2)
                else:
                    return f'HTTPS {response.status_code}', round(response_time, 2)
            except Exception:
                return 'Connection Error', None
        except Exception as e:
            return f'Error: {str(e)[:50]}', None
    
    def determine_overall_status(self, result: Dict) -> str:
        """Determine overall status based on probe results"""
        readiness = result['readiness_status']
        liveness = result['liveness_status']
        
        if readiness == 'Healthy' and liveness == 'Healthy':
            return 'Healthy'
        elif readiness == 'Not Configured' and liveness == 'Not Configured':
            return 'No Probes'
        elif readiness in ['Healthy', 'Not Configured'] and liveness in ['Healthy', 'Not Configured']:
            # At least one is healthy and none are unhealthy
            return 'Partially Healthy'
        elif readiness.startswith('Error') or liveness.startswith('Error'):
            return 'Error'
        elif 'Timeout' in readiness or 'Timeout' in liveness:
            return 'Timeout'
        elif 'Connection Error' in readiness or 'Connection Error' in liveness:
            return 'Connection Error'
        else:
            return 'Unhealthy'
    
    def generate_report(self) -> str:
        """Generate comprehensive report"""
        logger.info("Generating reports...")
        
        # Create reports directory with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        reports_dir = f"reports/{timestamp}"
        os.makedirs(reports_dir, exist_ok=True)
        
        # Generate detailed CSV report
        csv_filename = os.path.join(reports_dir, f"service_health_detailed.csv")
        self.generate_csv_report(csv_filename)
        
        # Generate Excel report with multiple sheets
        excel_filename = os.path.join(reports_dir, f"service_health_report.xlsx")
        self.generate_excel_report(excel_filename)
        
        # Generate summary report
        summary_filename = os.path.join(reports_dir, f"service_summary.json")
        self.generate_summary_report(summary_filename)
        
        # Generate HTML dashboard
        html_filename = os.path.join(reports_dir, f"service_dashboard.html")
        self.generate_html_dashboard(html_filename)
        
        logger.info(f"Reports generated in {reports_dir}")
        
        # Create latest symlink
        try:
            if os.path.exists("reports/latest"):
                os.unlink("reports/latest")
            os.symlink(timestamp, "reports/latest")
        except:
            pass  # Ignore symlink errors on Windows
        
        return excel_filename
    
    def generate_csv_report(self, filename: str):
        """Generate CSV report"""
        df = pd.DataFrame(self.results)
        
        # Reorder columns for better readability
        column_order = [
            'timestamp', 'cluster', 'namespace', 'service', 'service_type',
            'overall_status', 'readiness_status', 'liveness_status',
            'readiness_response_time', 'liveness_response_time',
            'pod_ip', 'node_name', 'readiness_path', 'liveness_path',
            'readiness_port', 'liveness_port', 'error'
        ]
        
        # Keep only columns that exist
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        df.to_csv(filename, index=False)
        logger.info(f"CSV report saved: {filename}")
    
    def generate_excel_report(self, filename: str):
        """Generate Excel report with multiple sheets and formatting"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.formatting.rule import CellIsRule, FormulaRule
            
            workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                default_sheet = workbook['Sheet']
                workbook.remove(default_sheet)
            
            # Summary sheet
            summary_df = pd.DataFrame(self.create_summary_data())
            self.add_sheet_with_formatting(workbook, 'Summary', summary_df)
            
            # Detailed results
            detailed_df = pd.DataFrame(self.results)
            self.add_sheet_with_formatting(workbook, 'Detailed Results', detailed_df)
            
            # Failed services
            if self.failed_services:
                failed_df = pd.DataFrame(self.failed_services)
                self.add_sheet_with_formatting(workbook, 'Failed Services', failed_df)
            
            # Cluster-wise summary
            cluster_summary = self.create_cluster_summary()
            cluster_df = pd.DataFrame(cluster_summary)
            self.add_sheet_with_formatting(workbook, 'Cluster Summary', cluster_df)
            
            # Namespace-wise summary
            namespace_summary = self.create_namespace_summary()
            namespace_df = pd.DataFrame(namespace_summary)
            self.add_sheet_with_formatting(workbook, 'Namespace Summary', namespace_df)
            
            # Save workbook
            workbook.save(filename)
            logger.info(f"Excel report saved: {filename}")
            
        except ImportError:
            logger.warning("openpyxl not installed. Generating simplified Excel report with pandas.")
            with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
                summary_df = pd.DataFrame(self.create_summary_data())
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                detailed_df = pd.DataFrame(self.results)
                detailed_df.to_excel(writer, sheet_name='Detailed Results', index=False)
                
                if self.failed_services:
                    failed_df = pd.DataFrame(self.failed_services)
                    failed_df.to_excel(writer, sheet_name='Failed Services', index=False)
                
                cluster_df = pd.DataFrame(self.create_cluster_summary())
                cluster_df.to_excel(writer, sheet_name='Cluster Summary', index=False)
    
    def add_sheet_with_formatting(self, workbook, sheet_name, df):
        """Add sheet with formatting to workbook"""
        from openpyxl.utils import get_column_letter
        
        # Create sheet
        ws = workbook.create_sheet(title=sheet_name)
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f"{col_letter}1"]
            cell.value = column_title
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_num, row in enumerate(df.itertuples(index=False), 2):
            for col_num, value in enumerate(row, 1):
                col_letter = get_column_letter(col_num)
                cell = ws[f"{col_letter}{row_num}"]
                cell.value = value
                
                # Apply conditional formatting for status columns
                if 'status' in str(value).lower() or 'healthy' in str(value).lower():
                    if 'unhealthy' in str(value).lower() or 'error' in str(value).lower():
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif 'healthy' in str(value).lower():
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_data(self) -> List[Dict]:
        """Create summary data for reporting"""
        total_services = len(self.results)
        healthy = len([r for r in self.results if r['overall_status'] == 'Healthy'])
        unhealthy = len([r for r in self.results if r['overall_status'] in ['Unhealthy', 'Error']])
        timeout = len([r for r in self.results if 'Timeout' in r['overall_status']])
        connection_error = len([r for r in self.results if 'Connection Error' in r['overall_status']])
        no_probes = len([r for r in self.results if r['overall_status'] == 'No Probes'])
        partially_healthy = len([r for r in self.results if r['overall_status'] == 'Partially Healthy'])
        
        health_score = (healthy/total_services*100) if total_services > 0 else 0
        
        summary = [
            {'Metric': 'Total Services Checked', 'Value': total_services, 'Details': ''},
            {'Metric': 'Healthy Services', 'Value': healthy, 
             'Details': f"{(healthy/total_services*100):.1f}%" if total_services > 0 else 'N/A'},
            {'Metric': 'Unhealthy/Error Services', 'Value': unhealthy,
             'Details': f"{(unhealthy/total_services*100):.1f}%" if total_services > 0 else 'N/A'},
            {'Metric': 'Timeout Services', 'Value': timeout,
             'Details': f"{(timeout/total_services*100):.1f}%" if total_services > 0 else 'N/A'},
            {'Metric': 'Connection Error Services', 'Value': connection_error,
             'Details': f"{(connection_error/total_services*100):.1f}%" if total_services > 0 else 'N/A'},
            {'Metric': 'Partially Healthy Services', 'Value': partially_healthy,
             'Details': f"{(partially_healthy/total_services*100):.1f}%" if total_services > 0 else 'N/A'},
            {'Metric': 'Services Without Probes', 'Value': no_probes,
             'Details': f"{(no_probes/total_services*100):.1f}%" if total_services > 0 else 'N/A'},
            {'Metric': 'Overall Health Score', 'Value': f"{health_score:.1f}%", 
             'Details': 'Excellent' if health_score >= 95 else 'Good' if health_score >= 85 else 'Needs Attention'}
        ]
        
        return summary
    
    def create_cluster_summary(self) -> List[Dict]:
        """Create cluster-wise summary"""
        cluster_summary = []
        
        for cluster in self.clusters:
            cluster_results = [r for r in self.results if r['cluster'] == cluster.name]
            total = len(cluster_results)
            
            if total == 0:
                continue
            
            healthy = len([r for r in cluster_results if r['overall_status'] == 'Healthy'])
            unhealthy = len([r for r in cluster_results if r['overall_status'] in ['Unhealthy', 'Error']])
            
            # Get namespace count for this cluster
            namespaces = set([r['namespace'] for r in cluster_results])
            
            cluster_summary.append({
                'Cluster': cluster.name,
                'Context': cluster.context,
                'Namespaces': len(namespaces),
                'Total Services': total,
                'Healthy': healthy,
                'Unhealthy': unhealthy,
                'Health %': f"{(healthy/total*100):.1f}%",
                'Status': '✅ Healthy' if (healthy/total*100) >= 95 else '⚠️ Needs Attention' if (healthy/total*100) >= 80 else '❌ Critical'
            })
        
        return cluster_summary
    
    def create_namespace_summary(self) -> List[Dict]:
        """Create namespace-wise summary"""
        namespace_data = {}
        
        for result in self.results:
            key = f"{result['cluster']}/{result['namespace']}"
            if key not in namespace_data:
                namespace_data[key] = {
                    'cluster': result['cluster'],
                    'namespace': result['namespace'],
                    'total': 0,
                    'healthy': 0,
                    'unhealthy': 0
                }
            
            namespace_data[key]['total'] += 1
            if result['overall_status'] == 'Healthy':
                namespace_data[key]['healthy'] += 1
            elif result['overall_status'] in ['Unhealthy', 'Error']:
                namespace_data[key]['unhealthy'] += 1
        
        namespace_summary = []
        for key, data in namespace_data.items():
            total = data['total']
            healthy = data['healthy']
            
            namespace_summary.append({
                'Cluster': data['cluster'],
                'Namespace': data['namespace'],
                'Total Services': total,
                'Healthy': healthy,
                'Unhealthy': data['unhealthy'],
                'Health %': f"{(healthy/total*100):.1f}%" if total > 0 else 'N/A',
                'Status': '✅' if (healthy/total*100) >= 95 else '⚠️' if (healthy/total*100) >= 80 else '❌'
            })
        
        return sorted(namespace_summary, key=lambda x: (x['Cluster'], x['Namespace']))
    
    def generate_summary_report(self, filename: str):
        """Generate JSON summary report"""
        summary = {
            'timestamp': datetime.now().isoformat(),
            'total_services': len(self.results),
            'healthy_services': len([r for r in self.results if r['overall_status'] == 'Healthy']),
            'failed_services': len(self.failed_services),
            'health_score': (len([r for r in self.results if r['overall_status'] == 'Healthy']) / len(self.results) * 100) if self.results else 0,
            'clusters_checked': [cluster.name for cluster in self.clusters],
            'cluster_summary': self.create_cluster_summary(),
            'top_failed_services': self.failed_services[:20]  # Limit to top 20
        }
        
        with open(filename, 'w') as f:
            json.dump(summary, f, indent=2, default=str)
        
        logger.info(f"Summary report saved: {filename}")
    
    def generate_html_dashboard(self, filename: str):
        """Generate HTML dashboard for visualization"""
        total = len(self.results)
        healthy = len([r for r in self.results if r['overall_status'] == 'Healthy'])
        failed = len(self.failed_services)
        health_score = (healthy/total*100) if total > 0 else 0
        
        # Determine status
        if health_score >= 95:
            status_color = "#28a745"
            status_text = "EXCELLENT"
        elif health_score >= 85:
            status_color = "#ffc107"
            status_text = "GOOD"
        else:
            status_color = "#dc3545"
            status_text = "NEEDS ATTENTION"
        
        # Create cluster charts data
        cluster_data = self.create_cluster_summary()
        cluster_chart_data = []
        for cluster in cluster_data:
            cluster_chart_data.append({
                'cluster': cluster['Cluster'],
                'healthy': cluster['Healthy'],
                'unhealthy': cluster['Unhealthy']
            })
        
        html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Kubernetes Service Health Dashboard</title>
            <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/luxon"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-luxon"></script>
            <style>
                * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f5f7fa; color: #333; line-height: 1.6; }}
                .container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px; }}
                .header h1 {{ font-size: 2.5em; margin-bottom: 10px; }}
                .header p {{ opacity: 0.9; }}
                .status-card {{ background: white; border-radius: 10px; padding: 25px; margin-bottom: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
                .status-indicator {{ display: inline-block; padding: 10px 25px; background: {status_color}; color: white; border-radius: 5px; font-weight: bold; margin-top: 10px; }}
                .metrics-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px; }}
                .metric-card {{ background: white; padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                .metric-value {{ font-size: 2.5em; font-weight: bold; margin: 10px 0; }}
                .metric-label {{ color: #666; text-transform: uppercase; font-size: 0.9em; }}
                .charts-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(500px, 1fr)); gap: 30px; margin-bottom: 30px; }}
                .chart-container {{ background: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                th, td {{ padding: 15px; text-align: left; border-bottom: 1px solid #eee; }}
                th {{ background: #f8f9fa; font-weight: 600; }}
                tr:hover {{ background: #f8f9fa; }}
                .healthy {{ color: #28a745; }}
                .unhealthy {{ color: #dc3545; }}
                .warning {{ color: #ffc107; }}
                .footer {{ margin-top: 40px; text-align: center; color: #666; font-size: 0.9em; }}
                .last-updated {{ text-align: right; color: #666; font-size: 0.9em; margin-bottom: 10px; }}
                @media (max-width: 768px) {{ .charts-grid {{ grid-template-columns: 1fr; }} }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🚀 Kubernetes Service Health Dashboard</h1>
                    <p>Automated Health Monitoring System | {datetime.now().strftime('%B %d, %Y %H:%M:%S UTC')}</p>
                </div>
                
                <div class="status-card">
                    <h2>Overall System Status</h2>
                    <div class="status-indicator">{status_text}</div>
                    <p>Health Score: <strong>{health_score:.1f}%</strong></p>
                </div>
                
                <div class="metrics-grid">
                    <div class="metric-card">
                        <div class="metric-label">Total Services</div>
                        <div class="metric-value">{total}</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-label">Healthy</div>
                        <div class="metric-value healthy">{healthy}</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-label">Failed</div>
                        <div class="metric-value unhealthy">{failed}</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-label">Health Score</div>
                        <div class="metric-value">{health_score:.1f}%</div>
                    </div>
                </div>
                
                <div class="charts-grid">
                    <div class="chart-container">
                        <h3>Cluster Health Distribution</h3>
                        <canvas id="clusterChart" height="300"></canvas>
                    </div>
                    <div class="chart-container">
                        <h3>Service Status Breakdown</h3>
                        <canvas id="statusChart" height="300"></canvas>
                    </div>
                </div>
                
                <div class="chart-container" style="margin-bottom: 30px;">
                    <h3>Failed Services</h3>
                    <table id="failedTable">
                        <thead>
                            <tr>
                                <th>Cluster</th>
                                <th>Namespace</th>
                                <th>Service</th>
                                <th>Type</th>
                                <th>Status</th>
                                <th>Readiness</th>
                                <th>Liveness</th>
                                <th>Error</th>
                            </tr>
                        </thead>
                        <tbody>
                            {"".join([f'''
                            <tr>
                                <td>{service['cluster']}</td>
                                <td>{service['namespace']}</td>
                                <td>{service['service']}</td>
                                <td>{service.get('service_type', 'N/A')}</td>
                                <td class="unhealthy">{service['overall_status']}</td>
                                <td>{service.get('readiness_status', 'N/A')}</td>
                                <td>{service.get('liveness_status', 'N/A')}</td>
                                <td title="{service.get('error', '')}">{service.get('error', 'N/A')[:50]}{'...' if service.get('error') and len(service.get('error', '')) > 50 else ''}</td>
                            </tr>
                            ''' for service in self.failed_services[:50]])}
                        </tbody>
                    </table>
                    {f'<p style="margin-top: 10px; color: #666;">Showing 50 of {len(self.failed_services)} failed services. Download full report for complete list.</p>' if len(self.failed_services) > 50 else ''}
                </div>
                
                <div class="footer">
                    <p>Generated by Kubernetes Service Monitor v2.0 | {datetime.now().year} Your Company</p>
                    <p>For issues or questions, contact: devops-team@yourcompany.com</p>
                </div>
            </div>
            
            <script>
                // Cluster Health Chart
                const clusterCtx = document.getElementById('clusterChart').getContext('2d');
                const clusterChart = new Chart(clusterCtx, {{
                    type: 'bar',
                    data: {{
                        labels: {json.dumps([c['Cluster'] for c in cluster_data])},
                        datasets: [
                            {{
                                label: 'Healthy',
                                data: {json.dumps([c['Healthy'] for c in cluster_data])},
                                backgroundColor: '#28a745'
                            }},
                            {{
                                label: 'Unhealthy',
                                data: {json.dumps([c['Unhealthy'] for c in cluster_data])},
                                backgroundColor: '#dc3545'
                            }}
                        ]
                    }},
                    options: {{
                        responsive: true,
                        plugins: {{
                            legend: {{ position: 'top' }},
                            title: {{ display: true, text: 'Services by Cluster' }}
                        }},
                        scales: {{
                            x: {{ stacked: true }},
                            y: {{ 
                                stacked: true,
                                beginAtZero: true,
                                title: {{ display: true, text: 'Number of Services' }}
                            }}
                        }}
                    }}
                }});
                
                // Status Breakdown Chart
                const statusCtx = document.getElementById('statusChart').getContext('2d');
                const healthyCount = {healthy};
                const unhealthyCount = {len([r for r in self.results if r['overall_status'] in ['Unhealthy', 'Error']])};
                const timeoutCount = {len([r for r in self.results if 'Timeout' in r['overall_status']])};
                const connectionErrorCount = {len([r for r in self.results if 'Connection Error' in r['overall_status']])};
                const partiallyHealthyCount = {len([r for r in self.results if r['overall_status'] == 'Partially Healthy'])};
                const noProbesCount = {len([r for r in self.results if r['overall_status'] == 'No Probes'])};
                
                const statusChart = new Chart(statusCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: ['Healthy', 'Unhealthy/Error', 'Timeout', 'Connection Error', 'Partially Healthy', 'No Probes'],
                        datasets: [{{
                            data: [healthyCount, unhealthyCount, timeoutCount, connectionErrorCount, partiallyHealthyCount, noProbesCount],
                            backgroundColor: [
                                '#28a745',
                                '#dc3545',
                                '#fd7e14',
                                '#6c757d',
                                '#ffc107',
                                '#17a2b8'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        plugins: {{
                            legend: {{ position: 'right' }},
                            title: {{ display: true, text: 'Service Status Distribution' }}
                        }}
                    }}
                }});
                
                // Auto-refresh page every 5 minutes
                setTimeout(() => {{
                    location.reload();
                }}, 300000);
            </script>
        </body>
        </html>
        """
        
        with open(filename, 'w') as f:
            f.write(html)
        
        logger.info(f"HTML dashboard saved: {filename}")
    
    def send_email_alert(self, report_file: str):
        """Send email alert with professional template"""
        logger.info("Preparing email alert...")
        
        if not self.email_config['sender_password']:
            logger.warning("Email password not configured. Skipping email alert.")
            return
        
        try:
            # Create message
            msg = MIMEMultipart('alternative')
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            msg['Subject'] = f'🚨 Kubernetes Service Health Alert - {timestamp}'
            msg['From'] = self.email_config['sender_email']
            msg['To'] = ', '.join(self.email_config['recipients'])
            if self.email_config['cc_recipients']:
                msg['Cc'] = ', '.join(self.email_config['cc_recipients'])
            
            # Create HTML email content
            html_content = self.create_email_html()
            msg.attach(MIMEText(html_content, 'html'))
            
            # Attach report files
            report_dir = os.path.dirname(report_file)
            for filename in os.listdir(report_dir):
                if filename.endswith(('.csv', '.xlsx', '.json')):
                    filepath = os.path.join(report_dir, filename)
                    with open(filepath, 'rb') as f:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(f.read())
                        encoders.encode_base64(part)
                        part.add_header('Content-Disposition', 
                                      f'attachment; filename="{filename}"')
                        msg.attach(part)
            
            # Send email
            all_recipients = self.email_config['recipients'] + self.email_config['cc_recipients']
            
            with smtplib.SMTP(self.email_config['smtp_server'], self.email_config['smtp_port']) as server:
                server.starttls()
                server.login(self.email_config['sender_email'], self.email_config['sender_password'])
                server.send_message(msg)
            
            logger.info(f"Email alert sent to {len(all_recipients)} recipients")
            
        except Exception as e:
            logger.error(f"Error sending email: {str(e)}")
    
    def create_email_html(self) -> str:
        """Create professional HTML email template"""
        
        total = len(self.results)
        healthy = len([r for r in self.results if r['overall_status'] == 'Healthy'])
        failed = len(self.failed_services)
        health_score = (healthy/total*100) if total > 0 else 0
        
        # Determine status color
        if health_score >= 95:
            status_color = "#28a745"  # Green
            status_text = "EXCELLENT"
            status_emoji = "✅"
        elif health_score >= 85:
            status_color = "#ffc107"  # Yellow
            status_text = "GOOD"
            status_emoji = "⚠️"
        else:
            status_color = "#dc3545"  # Red
            status_text = "ATTENTION NEEDED"
            status_emoji = "🚨"
        
        # Create cluster status table
        cluster_table = ""
        cluster_data = self.create_cluster_summary()
        for cluster in cluster_data:
            cluster_table += f"""
            <tr>
                <td style="padding: 12px; border-bottom: 1px solid #e0e0e0;">{cluster['Cluster']}</td>
                <td style="padding: 12px; border-bottom: 1px solid #e0e0e0; text-align: center;">{cluster['Total Services']}</td>
                <td style="padding: 12px; border-bottom: 1px solid #e0e0e0; text-align: center; color: #28a745;">{cluster['Healthy']}</td>
                <td style="padding: 12px; border-bottom: 1px solid #e0e0e0; text-align: center; color: #dc3545;">{cluster['Unhealthy']}</td>
                <td style="padding: 12px; border-bottom: 1px solid #e0e0e0; text-align: center;">{cluster['Health %']}</td>
                <td style="padding: 12px; border-bottom: 1px solid #e0e0e0; text-align: center;">{cluster['Status']}</td>
            </tr>
            """
        
        # Create failed services table (top 10)
        failed_table = ""
        for service in self.failed_services[:10]:
            error = service.get('error', 'N/A')
            if len(error) > 50:
                error = error[:50] + '...'
            
            failed_table += f"""
            <tr>
                <td style="padding: 10px; border-bottom: 1px solid #e0e0e0;">{service['cluster']}</td>
                <td style="padding: 10px; border-bottom: 1px solid #e0e0e0;">{service['namespace']}</td>
                <td style="padding: 10px; border-bottom: 1px solid #e0e0e0; font-weight: 500;">{service['service']}</td>
                <td style="padding: 10px; border-bottom: 1px solid #e0e0e0;">{service.get('service_type', 'N/A')}</td>
                <td style="padding: 10px; border-bottom: 1px solid #e0e0e0; color: #dc3545; font-weight: 500;">{service['overall_status']}</td>
                <td style="padding: 10px; border-bottom: 1px solid #e0e0e0;">{error}</td>
            </tr>
            """
        
        html_template = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Kubernetes Health Report</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 0; background-color: #f8f9fa; }}
                .email-container {{ max-width: 800px; margin: 0 auto; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.1); }}
                .email-header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 40px 30px; text-align: center; }}
                .email-header h1 {{ margin: 0; font-size: 28px; font-weight: 600; }}
                .email-header p {{ margin: 10px 0 0; opacity: 0.9; font-size: 16px; }}
                .email-content {{ padding: 30px; }}
                .status-banner {{ background: {status_color}; color: white; padding: 20px; border-radius: 8px; text-align: center; margin-bottom: 30px; }}
                .status-banner h2 {{ margin: 0; font-size: 24px; }}
                .status-banner p {{ margin: 10px 0 0; font-size: 18px; opacity: 0.9; }}
                .metrics-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 30px; }}
                .metric-box {{ background: #f8f9fa; border-radius: 8px; padding: 20px; text-align: center; border: 1px solid #e9ecef; }}
                .metric-value {{ font-size: 28px; font-weight: 700; margin: 5px 0; }}
                .metric-label {{ font-size: 12px; color: #6c757d; text-transform: uppercase; letter-spacing: 0.5px; }}
                .healthy {{ color: #28a745; }}
                .unhealthy {{ color: #dc3545; }}
                .section-title {{ font-size: 20px; font-weight: 600; margin: 30px 0 15px; color: #343a40; border-bottom: 2px solid #e9ecef; padding-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th {{ background-color: #f8f9fa; font-weight: 600; text-align: left; padding: 12px; border-bottom: 2px solid #e9ecef; color: #495057; }}
                td {{ padding: 12px; border-bottom: 1px solid #e9ecef; }}
                tr:hover {{ background-color: #f8f9fa; }}
                .email-footer {{ background: #f8f9fa; padding: 25px 30px; text-align: center; border-top: 1px solid #e9ecef; color: #6c757d; font-size: 14px; }}
                .email-footer a {{ color: #667eea; text-decoration: none; }}
                .attachments {{ background: #f8f9fa; padding: 15px; border-radius: 8px; margin-top: 20px; font-size: 14px; }}
                .critical-note {{ background: #fff3cd; border: 1px solid #ffc107; color: #856404; padding: 15px; border-radius: 8px; margin-top: 20px; }}
                @media (max-width: 600px) {{
                    .metrics-grid {{ grid-template-columns: repeat(2, 1fr); }}
                    table {{ font-size: 14px; }}
                }}
            </style>
        </head>
        <body>
            <div class="email-container">
                <div class="email-header">
                    <h1>🚀 Kubernetes Service Health Report</h1>
                    <p>Automated Monitoring System | {datetime.now().strftime('%B %d, %Y %H:%M UTC')}</p>
                </div>
                
                <div class="email-content">
                    <div class="status-banner">
                        <h2>{status_emoji} Overall Status: {status_text}</h2>
                        <p>Health Score: <strong>{health_score:.1f}%</strong></p>
                    </div>
                    
                    <div class="metrics-grid">
                        <div class="metric-box">
                            <div class="metric-label">Total Services</div>
                            <div class="metric-value">{total}</div>
                        </div>
                        <div class="metric-box">
                            <div class="metric-label">Healthy</div>
                            <div class="metric-value healthy">{healthy}</div>
                        </div>
                        <div class="metric-box">
                            <div class="metric-label">Failed</div>
                            <div class="metric-value unhealthy">{failed}</div>
                        </div>
                        <div class="metric-box">
                            <div class="metric-label">Health Score</div>
                            <div class="metric-value">{health_score:.1f}%</div>
                        </div>
                    </div>
                    
                    <div class="section-title">📊 Cluster-wise Summary</div>
                    <table>
                        <thead>
                            <tr>
                                <th>Cluster</th>
                                <th style="text-align: center;">Total</th>
                                <th style="text-align: center;">Healthy</th>
                                <th style="text-align: center;">Unhealthy</th>
                                <th style="text-align: center;">Health %</th>
                                <th style="text-align: center;">Status</th>
                            </tr>
                        </thead>
                        <tbody>
                            {cluster_table}
                        </tbody>
                    </table>
                    
                    {f'''
                    <div class="section-title">⚠️ Failed Services ({failed} total)</div>
                    <table>
                        <thead>
                            <tr>
                                <th>Cluster</th>
                                <th>Namespace</th>
                                <th>Service</th>
                                <th>Type</th>
                                <th>Status</th>
                                <th>Error</th>
                            </tr>
                        </thead>
                        <tbody>
                            {failed_table}
                        </tbody>
                    </table>
                    {f'<p style="color: #6c757d; font-size: 14px; margin-top: 10px;">Showing first 10 of {failed} failed services. See attached report for complete list.</p>' if failed > 10 else ''}
                    ''' if failed > 0 else '<div class="section-title">✅ All Services Healthy</div><p>No failed services detected.</p>'}
                    
                    {f'<div class="critical-note"><strong>⚠️ Attention Required:</strong> {failed} services are failing. Immediate action is recommended.</div>' if failed > 0 else ''}
                    
                    <div class="attachments">
                        <strong>📎 Attachments:</strong>
                        <ul style="margin: 10px 0 0; padding-left: 20px;">
                            <li>Detailed Service Health Report (Excel)</li>
                            <li>Service Health Data (CSV)</li>
                            <li>Summary Report (JSON)</li>
                            <li>Interactive Dashboard (HTML)</li>
                        </ul>
                    </div>
                </div>
                
                <div class="email-footer">
                    <p>This is an automated report generated by Kubernetes Service Monitor v2.0</p>
                    <p>For any questions or concerns, please contact the DevOps team at <a href="mailto:devops-team@yourcompany.com">devops-team@yourcompany.com</a></p>
                    <p style="margin-top: 15px; font-size: 12px; color: #adb5bd;">
                        © {datetime.now().year} Your Company Name. All rights reserved.<br>
                        This email was automatically generated. Please do not reply.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html_template
    
    def run(self):
        """Main execution method"""
        logger.info("=" * 80)
        logger.info("Starting Kubernetes Service Health Monitoring...")
        logger.info("=" * 80)
        
        try:
            start_time = datetime.now()
            
            # Step 1: Discover service probes dynamically
            self.discover_service_probes()
            
            if not self.service_probes:
                logger.error("No services with probes found. Exiting.")
                return
            
            # Step 2: Check all probes in parallel
            self.check_service_probes()
            
            # Step 3: Generate comprehensive reports
            report_file = self.generate_report()
            
            # Step 4: Send email alert (if configured)
            if self.email_config['sender_password']:
                self.send_email_alert(report_file)
            else:
                logger.warning("Email password not configured. Skipping email alert.")
            
            # Step 5: Print summary to console
            self.print_summary()
            
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()
            
            logger.info("=" * 80)
            logger.info(f"Monitoring completed successfully in {duration:.2f} seconds!")
            logger.info("=" * 80)
            
        except KeyboardInterrupt:
            logger.info("Monitoring interrupted by user")
        except Exception as e:
            logger.error(f"Monitoring failed: {str(e)}", exc_info=True)
            sys.exit(1)
    
    def print_summary(self):
        """Print summary to console"""
        print("\n" + "=" * 100)
        print("KUBERNETES SERVICE HEALTH MONITORING - EXECUTIVE SUMMARY")
        print("=" * 100)
        
        total = len(self.results)
        healthy = len([r for r in self.results if r['overall_status'] == 'Healthy'])
        failed = len(self.failed_services)
        
        print(f"\n📊 OVERALL STATISTICS:")
        print(f"   Total Services Checked: {total}")
        print(f"   ✅ Healthy Services: {healthy}")
        print(f"   ❌ Failed Services: {failed}")
        if total > 0:
            print(f"   📈 Health Score: {(healthy/total*100):.1f}%")
        
        print(f"\n🏢 CLUSTER-WISE BREAKDOWN:")
        cluster_data = self.create_cluster_summary()
        for cluster in cluster_data:
            status_icon = "✅" if "Healthy" in cluster['Status'] else "⚠️" if "Attention" in cluster['Status'] else "❌"
            print(f"   {status_icon} {cluster['Cluster']}: {cluster['Healthy']}/{cluster['Total Services']} "
                  f"healthy ({cluster['Health %']})")
        
        if self.failed_services:
            print(f"\n⚠️  TOP FAILED SERVICES ({min(5, len(self.failed_services))} shown):")
            for service in self.failed_services[:5]:
                print(f"   • {service['cluster']}/{service['namespace']}/{service['service']}")
                print(f"     Status: {service['overall_status']}")
                if service.get('error'):
                    print(f"     Error: {service['error'][:80]}...")
                print()
        
        print(f"\n📁 REPORTS:")
        print(f"   Detailed reports generated in: reports/latest/")
        print(f"   - Excel Report: service_health_report.xlsx")
        print(f"   - CSV Data: service_health_detailed.csv")
        print(f"   - Summary: service_summary.json")
        print(f"   - Dashboard: service_dashboard.html")
        
        print(f"\n⏰ NEXT STEPS:")
        if failed > 0:
            print(f"   ❗ {failed} services require immediate attention")
            print(f"   🔍 Check the failed services report for details")
        else:
            print(f"   ✅ All services are healthy. No action required.")
        
        print("=" * 100)


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Kubernetes Service Health Monitoring System',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --config custom_config.yaml
  %(prog)s --create-config
  %(prog)s --setup
  %(prog)s --dry-run
  
Required Setup:
  1. Install dependencies: pip install kubernetes requests pandas openpyxl pyyaml
  2. Configure kubectl contexts for all clusters
  3. Set email environment variables (optional)
        """
    )
    
    parser.add_argument('--config', default='cluster_config.yaml', 
                       help='Path to configuration file (default: cluster_config.yaml)')
    parser.add_argument('--create-config', action='store_true', 
                       help='Create sample configuration file')
    parser.add_argument('--setup', action='store_true', 
                       help='Show setup instructions')
    parser.add_argument('--dry-run', action='store_true',
                       help='Discover services without checking probes')
    parser.add_argument('--verbose', '-v', action='store_true',
                       help='Enable verbose logging')
    parser.add_argument('--no-email', action='store_true',
                       help='Disable email alerts')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.info("Verbose logging enabled")
    
    if args.create_config:
        monitor = KubernetesServiceMonitor()
        monitor.create_default_config()
        print("✅ Sample configuration file created: cluster_config.yaml")
        print("   Please update the contexts according to your environment.")
        return
    
    if args.setup:
        print_setup_instructions()
        return
    
    # Run the monitor
    monitor = KubernetesServiceMonitor(args.config)
    
    if args.dry_run:
        logger.info("Running in dry-run mode (discovery only)")
        monitor.discover_service_probes()
        print(f"\nDiscovered {len(monitor.service_probes)} services with probes:")
        for probe in monitor.service_probes[:20]:  # Show first 20
            print(f"  - {probe.cluster}/{probe.namespace}/{probe.service_name}")
        if len(monitor.service_probes) > 20:
            print(f"  ... and {len(monitor.service_probes) - 20} more")
        return
    
    if args.no_email:
        monitor.email_config['sender_password'] = ''
        logger.info("Email alerts disabled")
    
    monitor.run()


def print_setup_instructions():
    """Print setup instructions"""
    print("\n" + "=" * 80)
    print("KUBERNETES SERVICE MONITOR - SETUP INSTRUCTIONS")
    print("=" * 80)
    
    print("\n1. INSTALL DEPENDENCIES:")
    print("   pip install kubernetes requests pandas openpyxl pyyaml")
    
    print("\n2. CONFIGURE KUBECTL CONTEXTS:")
    print("   Ensure all cluster contexts are configured in your kubeconfig:")
    print("   kubectl config get-contexts")
    print("   ")
    print("   Required contexts:")
    print("     - uat-general")
    print("     - uatbz-general")
    print("     - uatbz-ai")
    print("     - uatbz-bot")
    print("     - preprod-general")
    print("     - preprod-ai")
    print("     - preprod-bot")
    
    print("\n3. CONFIGURE EMAIL (Optional):")
    print("   export SMTP_SERVER='smtp.gmail.com'")
    print("   export SMTP_PORT='587'")
    print("   export SENDER_EMAIL='your-email@gmail.com'")
    print("   export SENDER_PASSWORD='your-app-password'")
    print("   export RECIPIENTS='team@company.com,manager@company.com'")
    
    print("\n4. CREATE CONFIGURATION FILE:")
    print("   python service_monitor.py --create-config")
    print("   Edit cluster_config.yaml to match your environment")
    
    print("\n5. RUN THE MONITOR:")
    print("   python service_monitor.py")
    print("   ")
    print("   Options:")
    print("     --config custom_config.yaml  # Use custom config")
    print("     --dry-run                    # Discovery only")
    print("     --no-email                   # Disable email")
    print("     --verbose                    # Enable debug logging")
    
    print("\n6. SCHEDULE AUTOMATED RUNS (Cron):")
    print("   */5 * * * * cd /path/to/monitor && python service_monitor.py")
    
    print("\n7. CHECK OUTPUT:")
    print("   Reports are saved in: reports/<timestamp>/")
    print("   Latest reports symlink: reports/latest")
    
    print("=" * 80)


if __name__ == "__main__":
    main()
